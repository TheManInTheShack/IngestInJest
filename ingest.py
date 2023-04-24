# ------------------------------------------------------------------------------
# Ingest
# Run a feedback loop for parsing raw text and supplementing it with information
# ------------------------------------------------------------------------------

# ------------------------------------------------------------------------------
# Imports
# ------------------------------------------------------------------------------
import sys
import argparse
import re
import os
import json

from datetime import datetime
from unidecode import unidecode
from nltk.corpus import wordnet as wn
from textblob import TextBlob

import openpyxl as xl
import pandas as pd

from dopes.io_tools import kill_program
from dopes.mapping_tools import read_map_sheet
from dopes.data_tools import matrixify, show_me
from dopes.text_tools import abstract_text
from dopes.excel_tools import can_write_to_excel, write_matrix_to_excel_sheet, write_excel_sheet, apply_formatting_to_cell

# ------------------------------------------------------------------------------
# Command line interface
# ------------------------------------------------------------------------------
def cli():
    parser = argparse.ArgumentParser()
    #parser.add_argument("some_arg_alas",  type=str, help="help text to be displayed")
    args = parser.parse_args()
    return args

# ------------------------------------------------------------------------------
# Initialization
# ------------------------------------------------------------------------------
def initialize(args):
    init = {}
    #init['rfile'] = os.path.join("data","raw_text.txt")
    init['rfile'] = os.path.join("data","raw_text_small.txt")
    init['cfile'] = os.path.join("data","cleaning.json")
    init['ifile'] = os.path.join("data","injest.xlsx")
    init['vdir']  = os.path.join("vault","Ingest InJest")
    return init

# ------------------------------------------------------------------------------
# Main
# ------------------------------------------------------------------------------
def main(args, init):
    # --------------------------------------------------------------------------
    # Start
    # --------------------------------------------------------------------------
    print("starting...")

    # --------------------------------------------------------------------------
    # Parse the spreadsheet data
    # --------------------------------------------------------------------------
    postag = read_map_sheet(init['ifile'],'pos_tags')['POS Tags']
    postag = postag.set_index('POS')
    data_snt = read_map_sheet(init['ifile'],'Sentences')['Sentences']
    data_snt = data_snt.set_index(['Chapter Idx','Paragraph Idx'])

    # --------------------------------------------------------------------------
    # Parse out an organized version of the raw book text
    # --------------------------------------------------------------------------
    corpus = parse_raw_text(init['rfile'], data_snt)

    # --------------------------------------------------------------------------
    # Read the vault files and pull out things we've marked there
    # --------------------------------------------------------------------------
    vault = parse_obsidian_vault(init['vdir'])

    # --------------------------------------------------------------------------
    # Compile the basic lexicon from the raw text and the spreadsheet
    # --------------------------------------------------------------------------
    lex_data = compile_lexicon(corpus)

    # --------------------------------------------------------------------------
    # Write the lexical data to the grid in the spreadsheet
    # --------------------------------------------------------------------------
    o = write_output_page_lexicon(lex_data, init['ifile'])

    # --------------------------------------------------------------------------
    # Write the sheets for the levels in the spreadsheet
    # --------------------------------------------------------------------------
    o = write_output_page_chapter(lex_data, init['ifile'], corpus)
    o = write_output_page_paragraph(lex_data, init['ifile'], corpus)
    o = write_output_page_sentence(lex_data, init['ifile'], corpus)

    # --------------------------------------------------------------------------
    # Write text files to the vault
    # --------------------------------------------------------------------------
    #vault_data = compile_vault_output(corpus)
    #o = write_vault_files(vault_data)

    # --------------------------------------------------------------------------
    # Finish
    # --------------------------------------------------------------------------
    print("...finished!")

# ------------------------------------------------------------------------------
# Read the raw text and parse it out into components
# ------------------------------------------------------------------------------
def parse_raw_text(rfile, data_snt):
    # --------------------------------------------------------------------------
    # Start
    # --------------------------------------------------------------------------
    print("...reading raw file...")

    # --------------------------------------------------------------------------
    # Get the raw text 
    # --------------------------------------------------------------------------
    raw = open(rfile).read().splitlines()

    # --------------------------------------------------------------------------
    # Get the cleaning instructions
    # --------------------------------------------------------------------------
    cleaning = {}
    cleaning['toc']  = {}
    cleaning['text'] = {}
    cleaning['note'] = {}

    # --------------------------------------------------------------------------
    # Work through each line and classify it
    # --------------------------------------------------------------------------
    toc  = []
    text = {}
    note = {}
    chapnum = 0
    heading = -1
    current_section = "start"
    current_chapter  = ""
    current_toc = ()
    current_note = 0
    for i, line in enumerate(raw):
        # ----------------------------------------------------------------------
        # Trim the input line
        # ----------------------------------------------------------------------
        line = line.strip()

        # ----------------------------------------------------------------------
        # Headings separating the sections are in the raw text like this:
        #
        # =====================================
        # SECTION NAME
        # =====================================
        #
        # This swaps a flag when we hit one of the lines to say whether or not 
        # we are 'inside' a heading, then goes to the next line
        # ----------------------------------------------------------------------
        if line.startswith("======="):
            heading *= -1
            continue

        # ----------------------------------------------------------------------
        # Switch sections to say which is current, based on that heading name
        # ----------------------------------------------------------------------
        if heading:
            if current_section == "start" and line.upper() == "TABLE OF CONTENTS":
                current_section = "toc"
                continue
            elif current_section == "start":
                continue
            elif current_section == "toc" and line.upper() == "TEXT":
                chapters = [x[2].upper() for x in toc]
                current_section = "text"
                continue
            elif current_section == "text" and line.upper() == "NOTES":
                current_section = "notes"
                continue

        # ----------------------------------------------------------------------
        # Table of contents, each line is a chapter name and page
        # ----------------------------------------------------------------------
        if current_section == "toc":
            # ------------------------------------------------------------------
            # White space doesn't mean anything
            # ------------------------------------------------------------------
            if len(line) == 0:
                continue

            # ------------------------------------------------------------------
            # Pull the line apart to get the page number and title, and 
            # increment the chapter number
            # ------------------------------------------------------------------
            pnum = line.split(".")[-1]
            title = line.replace("."+pnum,"")

            chapnum += 1

            # ------------------------------------------------------------------
            # Set the information as a tuple so we can use it to index the text
            # ------------------------------------------------------------------
            chapter = (chapnum, int(pnum), title)

            # ------------------------------------------------------------------
            # Add to the structure
            # ------------------------------------------------------------------
            toc.append(chapter)
            text[chapter] = {}

        # ----------------------------------------------------------------------
        # Text, we want to track by chapter. Each line will be indexed as part
        # of a chapter-paragraph-sentence hierarchy, with separate indices for 
        # the page and the original line
        # ----------------------------------------------------------------------
        elif current_section == "text":
            # ------------------------------------------------------------------
            # Determine if we are at the start of a chapter. Since some 
            # chapters have the same names, we need to go by both the chapter
            # name alone, which we would see in the line, and the toc, which
            # gives us the meta-information for the chapter and is the match key
            # ------------------------------------------------------------------
            if len(chapters) > 0 and line.upper() == chapters[0]:
                # --------------------------------------------------------------
                # Get the chapter information and pull this one off the lists
                # to expose the next one
                # --------------------------------------------------------------
                current_chapter = chapters.pop(0)
                current_toc     = toc.pop(0)

                # --------------------------------------------------------------
                # Start the structures for the chapter
                # --------------------------------------------------------------
                text[current_toc]['lines'] = []
                para_num  = 0
                line_num  = 0
                chap_num  = current_toc[0]
                page_num  = int(current_toc[1])
                chap_name = current_toc[2]
                empty     = 0
                previous  = 0

                # --------------------------------------------------------------
                # We don't need to do anything else on this line
                # --------------------------------------------------------------
                continue

            # ------------------------------------------------------------------
            # If we haven't even gotten to the first chapter, skip
            # ------------------------------------------------------------------
            elif current_chapter == "":
                continue

            # ------------------------------------------------------------------
            # Three empty lines indicates a page break, one indicates paragraph
            # breaks...so we need to track the empty lines
            # ------------------------------------------------------------------
            if len(line) == 0:
                empty += 1
            else:
                empty = 0

            # ------------------------------------------------------------------
            # Flip the page number if it's been three
            # ------------------------------------------------------------------
            if empty == 3:
                page_num += 1

            # ------------------------------------------------------------------
            # If we're starting up new text after a paragraph break, flip it
            # ------------------------------------------------------------------
            if empty == 0 and previous == 1:
                para_num += 1

            # ------------------------------------------------------------------
            # Increment the line number, disregarding extra whitespace
            # ------------------------------------------------------------------
            if empty == 0 or empty == 1:
                line_num += 1

            # ------------------------------------------------------------------
            # Save the empty flag for the next line
            # ------------------------------------------------------------------
            previous = empty

            # ------------------------------------------------------------------
            # go ahead and dump the whitespace
            # ------------------------------------------------------------------
            if empty:
                continue

            # ------------------------------------------------------------------
            # TODO
            # Can we identify where the note references are?
            # ------------------------------------------------------------------

            # ------------------------------------------------------------------
            # Add the line to the structure
            # ------------------------------------------------------------------
            slug = (para_num, line_num, page_num, line)
            text[current_toc]['lines'].append(slug)

        # ----------------------------------------------------------------------
        # Notes are numbered sequentially, and may be a lot of text with bullet
        # points, etc.
        # ----------------------------------------------------------------------
        elif current_section == "notes":
            # ------------------------------------------------------------------
            # Identify the first line and start the section
            # ------------------------------------------------------------------
            if line.startswith(str(current_note+1)):
                current_note += 1
                note[current_note] = []

            # ------------------------------------------------------------------
            # skip out if we haven't started yet
            # ------------------------------------------------------------------
            if current_note == 0:
                continue

            # ------------------------------------------------------------------
            # Add the line to the appropriate note
            # ------------------------------------------------------------------
            note[current_note].append(line)

    # --------------------------------------------------------------------------
    # Make a pass through the text and compile the lines to paragraphs
    # --------------------------------------------------------------------------
    data_paragraph = {}
    for i, chapter in enumerate(text):
        # ----------------------------------------------------------------------
        # Initialize the paragraph holder
        # ----------------------------------------------------------------------
        text[chapter]['paragraphs'] = {}

        # ----------------------------------------------------------------------
        # Work through each line in the chapter
        # ----------------------------------------------------------------------
        for j, line in enumerate(text[chapter]['lines']):
            # ------------------------------------------------------------------
            # Make the index ID
            # ------------------------------------------------------------------
            para_idx = (chapter,line[0]+1)

            # ------------------------------------------------------------------
            # If the paragraph isn't there yet, start it 
            # ------------------------------------------------------------------
            if not para_idx in text[chapter]['paragraphs']:
                text[chapter]['paragraphs'][para_idx] = ""

            if not para_idx in data_paragraph:
                data_paragraph[para_idx] = {}

            # ------------------------------------------------------------------
            # Line transitions
            # ------------------------------------------------------------------
            if len(text[chapter]['paragraphs'][para_idx]) > 0:
                # --------------------------------------------------------------
                # The end of the previous line is a hyphen, so no space
                # --------------------------------------------------------------
                if text[chapter]['paragraphs'][para_idx].endswith("-"):
                    text[chapter]['paragraphs'][para_idx] += ""

                # --------------------------------------------------------------
                # Normal is space betwen words on lines
                # --------------------------------------------------------------
                else:
                    text[chapter]['paragraphs'][para_idx] += " "

            # ------------------------------------------------------------------
            # Add the new line contents
            # ------------------------------------------------------------------
            text[chapter]['paragraphs'][para_idx] += line[3]

    # --------------------------------------------------------------------------
    # Another pass to break the paragraphs into sentences and apply any cleaning
    # --------------------------------------------------------------------------
    for i, chapter in enumerate(text):
        # ----------------------------------------------------------------------
        # Initialize the sentence holder
        # ----------------------------------------------------------------------
        text[chapter]['sentences'] = {}

        # ----------------------------------------------------------------------
        # Work through each paragraph in the chapter
        # ----------------------------------------------------------------------
        for j, para_idx in enumerate(text[chapter]['paragraphs']):
            # ------------------------------------------------------------------
            # Break the paragraph into 'natural' sentences, meaning whatever the
            # parser gives us
            # ------------------------------------------------------------------
            tb = TextBlob(text[chapter]['paragraphs'][para_idx])
            sentences = tb.sentences

            # ------------------------------------------------------------------
            # Get the existing version from the data to look for cleaning marks
            # ------------------------------------------------------------------
            cidx = para_idx[0][0]
            pidx = para_idx[1]
            key = (cidx, pidx)
            sdata = data_snt.loc[key]
            sdata = sdata.reset_index()
            sdata = sdata[['Sentence Idx','Cleaning Instruction','Text']]

            # ------------------------------------------------------------------
            # Work through each sentence, maybe clean, add it to the structure
            # ------------------------------------------------------------------ 
            clean_sentences = []
            skips = 0
            for k, sentence in enumerate(sentences):
                # --------------------------------------------------------------
                # Maybe we'll be skipping the line because of an instruction on
                # a previous line
                # --------------------------------------------------------------
                if skips > 0:
                    skips -= 1
                    continue

                # --------------------------------------------------------------
                # Cast the sentence as just the string of itself to simplify
                # --------------------------------------------------------------
                sentence = str(sentence)

                # --------------------------------------------------------------
                # The sentence index will be used a bunch
                # --------------------------------------------------------------
                sidx = k+1

                # --------------------------------------------------------------
                # Maybe there's a new instruction, so grab whatever is in that
                # cell for this line.  It'll be semicolon separated, maybe.
                # --------------------------------------------------------------
                cleans = []
                if k in sdata.index.tolist():
                    cleans = [x.lower().strip().replace(" ","") for x in sdata.loc[k]['Cleaning Instruction'].split(";") if x]

                # --------------------------------------------------------------
                # Validate and refine the instructions
                # --------------------------------------------------------------
                for item in cleans:
                    # ----------------------------------------------------------
                    # Combine
                    # ----------------------------------------------------------
                    if item.startswith("combine"):
                        # ------------------------------------------------------
                        # What we really want is a number
                        # ------------------------------------------------------
                        combine_num = 0

                        # ------------------------------------------------------
                        # The instruction is at least written correctly
                        # ------------------------------------------------------
                        kill = False
                        if not "+" in item:
                            kill = True
                        pieces = item.split("+")
                        if not len(pieces) == 2:
                            kill = True
                        try:
                            combine_num = int(pieces[1])
                        except:
                            kill = True
                        if kill:
                            kill_program(f"You've got a bad combine instruction in the sentence cleaning column. It should read like 'combine+1', where the 1 is however many following lines.  Found at c/p/s address {cidx}, {pidx}, {k}")

                        # ------------------------------------------------------
                        # Makes sense in context with the paragraph
                        # ------------------------------------------------------
                        if sidx + combine_num > len(sentences):
                            kill_program(f"You've got a combine instruction in the sentence cleaning that would exceed the length of its paragraph.  Found at c/p/s address {cidx}, {pidx}, {k}")

                    # ----------------------------------------------------------
                    # Correct text
                    # ----------------------------------------------------------
                    if item.startswith("correct"):
                        #TODO
                        pass

                # --------------------------------------------------------------
                # Now we have a good list of cleans; add it back into the data
                # structure that will be saved back out
                # --------------------------------------------------------------
                if cleans:
                    if not cidx in cleaning['text']:
                        cleaning['text'][cidx] = {}
                    if not pidx in cleaning['text'][cidx]:
                        cleaning['text'][cidx][pidx] = {}
                    cleaning['text'][cidx][pidx][sidx] = cleans

                # --------------------------------------------------------------
                # Apply the cleaning
                # --------------------------------------------------------------
                for clean in cleans:
                    # ----------------------------------------------------------
                    # If it's a combo, rewrite the line as the combination of 
                    # itself and its marked followers; also here set that many
                    # of the next ones to skip
                    # ----------------------------------------------------------
                    if clean.startswith("combine"):
                        sent = sentence
                        cnum = int(clean.split("+")[1])
                        for l in range(cnum):
                            sent += " " + str(sentences[k+l+1])
                            skips = cnum
                        sentence = sent

                # --------------------------------------------------------------
                # Write clean sentence to list
                # --------------------------------------------------------------
                clean_sentences.append((sidx, sentence))

            # ------------------------------------------------------------------
            # Write the number of sentences
            # ------------------------------------------------------------------
            data_paragraph[para_idx]['num_sentences_orignal'] = len(sentences)
            data_paragraph[para_idx]['num_sentences_clean']   = len(clean_sentences)

            # ------------------------------------------------------------------
            # Add each sentence to the big text structure
            # ------------------------------------------------------------------
            for k, sent_data in enumerate(clean_sentences):
                # --------------------------------------------------------------
                # Make the index ID
                # --------------------------------------------------------------
                sent_idx = (para_idx, int(sent_data[0]), k+1)

                # --------------------------------------------------------------
                # Add the sentence to the raw text area
                # --------------------------------------------------------------
                text[chapter]['sentences'][sent_idx] = sent_data[1]

    # --------------------------------------------------------------------------
    # Set up the different levels that we want to lexify at by re-compiling the
    # text back up into a linear basic view for each
    # --------------------------------------------------------------------------
    text_chapter   = {}
    text_paragraph = {}
    text_sentence  = {}
    text_wordtags  = []
    for chapter in text:
        # ----------------------------------------------------------------------
        # Do the paragraph and compile the chapter as we go
        # ----------------------------------------------------------------------
        buff_chapter = ""
        for para_idx in text[chapter]['paragraphs']:
            para_text = text[chapter]['paragraphs'][para_idx]
            buff_chapter += " " + para_text
            text_paragraph[para_idx] = para_text

        # ----------------------------------------------------------------------
        # Write the chapter that got compiled
        # ----------------------------------------------------------------------
        text_chapter[chapter] = buff_chapter

        # ----------------------------------------------------------------------
        # Do the sentences - note that before we text-ified the sentence so that
        # we might combine it; now we have to re-TextBlob it to get the tags
        # ----------------------------------------------------------------------
        for sent_idx in text[chapter]['sentences']:
            sent_text = text[chapter]['sentences'][sent_idx]
            text_sentence[sent_idx] = sent_text

            for wordtag in TextBlob(sent_text).tags:
                text_wordtags.append(wordtag)

    # --------------------------------------------------------------------------
    # Compile the plain sentence text back up to the full volume
    # --------------------------------------------------------------------------
    text_volume = " ".join([text_sentence[x] for x in text_sentence])

    # --------------------------------------------------------------------------
    # User info
    # --------------------------------------------------------------------------
    print(f"......chapters:   {len(text_chapter)}")
    print(f"......paragraphs: {len(text_paragraph)}")
    print(f"......sentences:  {len(text_sentence)}")
    print(f"......words:      {len(text_wordtags)}")
    print(f"......glyphs:     {len(text_volume)}")

    # --------------------------------------------------------------------------
    # Consolidate
    # --------------------------------------------------------------------------
    corpus = {}
    corpus['toc']  = toc
    corpus['text'] = text
    corpus['note'] = note

    corpus['text_lvl'] = {}
    corpus['text_lvl']['volume']    = text_volume
    corpus['text_lvl']['chapter']   = text_chapter
    corpus['text_lvl']['paragraph'] = text_paragraph
    corpus['text_lvl']['sentence']  = text_sentence
    corpus['text_lvl']['word_tags'] = text_wordtags

    corpus['data_lvl'] = {}
    corpus['data_lvl']['paragraph'] = data_paragraph

    corpus['cleaning'] = cleaning

    # --------------------------------------------------------------------------
    # Finish
    # --------------------------------------------------------------------------
    return corpus

# ------------------------------------------------------------------------------
# Pull out all the text and organize the base-level lexicon stuff
# ------------------------------------------------------------------------------
def compile_lexicon(corpus):
    # --------------------------------------------------------------------------
    # Start
    # --------------------------------------------------------------------------
    print("...compiling lexicon from primary text...")

    # --------------------------------------------------------------------------
    # Do the lexing at each level
    # --------------------------------------------------------------------------
    print("......compiling volume level...")
    lex_volume = {}
    lex_volume['full_text'] = lexify_one_set(corpus['text_lvl']['volume'])

    print("......compiling chapter level...")
    lex_chapter = {}
    for idx in corpus['text_lvl']['chapter']:
        lex_chapter[idx] = lexify_one_set(corpus['text_lvl']['chapter'][idx])

    print("......compiling paragraph level...")
    lex_paragraph = {}
    for idx in corpus['text_lvl']['paragraph']:
        lex_paragraph[idx] = lexify_one_set(corpus['text_lvl']['paragraph'][idx])

    print("......compiling sentence level...")
    lex_sentence = {}
    for idx in corpus['text_lvl']['sentence']:
        lex_sentence[idx] = lexify_one_set(corpus['text_lvl']['sentence'][idx])

    #for idx in lex_sentence:
    #    print("----------------------------------------------")
    #    print(idx)
    #    print(corpus['text_lvl']['sentence'][idx])
    #    print(lex_sentence[idx]['counts'])
    #    print(lex_sentence[idx]['counts_words'])

    # --------------------------------------------------------------------------
    # Consolidate
    # --------------------------------------------------------------------------
    lex_data = {}
    lex_data['volume']    = lex_volume
    lex_data['chapter']   = lex_chapter
    lex_data['paragraph'] = lex_paragraph
    lex_data['sentence']  = lex_sentence

    # --------------------------------------------------------------------------
    # FInish
    # --------------------------------------------------------------------------
    return lex_data

# ------------------------------------------------------------------------------
# Given one block of text, lexify it at the word/part of speech level and return
# several useful views
# ------------------------------------------------------------------------------
def lexify_one_set(text):
    # --------------------------------------------------------------------------
    # Make a textblob out of the thing we are passed
    # --------------------------------------------------------------------------
    blob = TextBlob(text)

    # --------------------------------------------------------------------------
    # Pull a list of all words broken out by part of speech tags and let's
    # hold a big list of those
    # --------------------------------------------------------------------------
    wtags = []
    words = []
    tags = []
    for word in blob.tags:
        wtags.append((word))
        words.append(word[0])
        tags.append(word[1])

    # --------------------------------------------------------------------------
    # Make simple counts of the unique set of words with tags
    # --------------------------------------------------------------------------
    counts_wtags = pd.Series(wtags, dtype="object").value_counts()
    counts_words = pd.Series(words, dtype="str").value_counts()
    counts_tags  = pd.Series(tags,  dtype="str").value_counts()

    unique_wtags = counts_wtags.index.to_list()
    unique_words = sorted(counts_words.index.to_list())
    unique_tags  = sorted(counts_tags.index.to_list())

    # --------------------------------------------------------------------------
    # Go back through the unique words and compile a few lookups
    # --------------------------------------------------------------------------
    tags_by_word = {}
    words_by_tag = {}
    abstract     = {}
    for word_tag in unique_wtags:
        # -----------------------------------------------------------------------
        # Unpack
        # -----------------------------------------------------------------------
        word = word_tag[0]
        tag  = word_tag[1]

        # -----------------------------------------------------------------------
        # Do cross-lists of tags per word and words per tag
        # -----------------------------------------------------------------------
        if not word in tags_by_word:
            tags_by_word[word] = []
        if not tag in words_by_tag:
            words_by_tag[tag] = []

        if not tag in tags_by_word[word]:
            tags_by_word[word].append(tag)
        if not word in words_by_tag[tag]:
            words_by_tag[tag].append(word)

        # ----------------------------------------------------------------------
        # Get the abstracts of the words
        # ----------------------------------------------------------------------
        abstract[word] = abstract_text(word)

    # --------------------------------------------------------------------------
    # Consolidate
    # --------------------------------------------------------------------------
    lexicon = {}
    lexicon['wtags']        = wtags
    lexicon['unique_wtags'] = unique_wtags
    lexicon['counts_wtags'] = counts_wtags

    lexicon['words']        = words
    lexicon['unique_words'] = unique_words
    lexicon['counts_words'] = counts_words

    lexicon['pos_tags']    = tags
    lexicon['unique_tags'] = unique_tags
    lexicon['counts_tags'] = counts_tags

    lexicon['tags_by_word'] = tags_by_word
    lexicon['words_by_tag'] = words_by_tag

    lexicon['abstract']     = abstract

    lexicon['counts'] = {}
    lexicon['counts']['total'] = len(words)
    lexicon['counts']['words'] = len(unique_words)
    lexicon['counts']['tags']  = len(unique_tags)

    # --------------------------------------------------------------------------
    # Finish
    # --------------------------------------------------------------------------
    return lexicon

# ------------------------------------------------------------------------------
# Pull the files out of the vault and do a basic parse 
# ------------------------------------------------------------------------------
def parse_obsidian_vault(vdir):
    # --------------------------------------------------------------------------
    # Start
    # --------------------------------------------------------------------------
    print("...parsing Obsidian vault, looking for new feedback...")

    # --------------------------------------------------------------------------
    # TODO This goes here...
    # --------------------------------------------------------------------------
    vinfo = {}

    # --------------------------------------------------------------------------
    # Finish
    # --------------------------------------------------------------------------
    return vinfo

# ------------------------------------------------------------------------------
# Replace the lexicon page with the latest version
# ------------------------------------------------------------------------------
def write_output_page_lexicon(lex_data, ifile):
    # --------------------------------------------------------------------------
    # Start
    # --------------------------------------------------------------------------
    print("...writing lexicon data to spreadsheet...")

    # --------------------------------------------------------------------------
    # File has to be available for writing, i.e. exists and not currently open
    # --------------------------------------------------------------------------
    if not can_write_to_excel(ifile):
        kill_program(f"Not able to write to Excel file {ifile}")

    # --------------------------------------------------------------------------
    # We just want the volume level lexicon here
    # --------------------------------------------------------------------------
    lexicon = lex_data['volume']['full_text']

    # --------------------------------------------------------------------------
    # Contents
    # --------------------------------------------------------------------------
    header = []
    header.append(["Lexicon","","","","","","","","","User"])
    header.append([])
    header.append(['Word','total count','Penn-Treebank tags','blanched','chopped','folded','mashed','juiced','x'])

    body = []
    for word in lexicon['unique_words']:
        postaglbl = ",".join(lexicon['tags_by_word'][word])

        rec = []
        rec.append(word)
        rec.append(lexicon['counts_words'][word])
        rec.append(postaglbl)
        rec.append(lexicon['abstract'][word][0])
        rec.append(lexicon['abstract'][word][1])
        rec.append(lexicon['abstract'][word][2])
        rec.append(lexicon['abstract'][word][3])
        rec.append(lexicon['abstract'][word][4])
        rec.append("")

        body.append(rec)

    data = matrixify([header,body])

    # --------------------------------------------------------------------------
    # Formatting
    # --------------------------------------------------------------------------
    fmt = {}
    fmt['column_widths']  = [30, 10, 20, 30, 30, 30, 30, 30, 5]
    fmt['reverse_rows']   = [1,2,3]
    fmt['short_rows']     = [2]
    fmt['wrap_rows']      = [3]
    fmt['fix_rows_after'] = 3
    fmt['freeze_panes']   = "A4"
    fmt['gutter_cols']    = ["I"]
    fmt['zoom']           = 80

    # --------------------------------------------------------------------------
    # Write the sheet in the file
    # --------------------------------------------------------------------------
    wb = xl.load_workbook(ifile)
    o = write_excel_sheet(wb, "Lexicon", data, fmt)
    wb.save(ifile)
    wb.close()

    # --------------------------------------------------------------------------
    # Finish
    # --------------------------------------------------------------------------
    return True

# ------------------------------------------------------------------------------
# Write out one page to the spreadsheet for a particular 'level' of text
# ------------------------------------------------------------------------------
def write_output_page_chapter(lex_data, ifile, corpus):
    # --------------------------------------------------------------------------
    # Start
    # --------------------------------------------------------------------------
    print(f"...writing chapter data to spreadsheet...")

    # --------------------------------------------------------------------------
    # Get the lexicon data for this particular level
    # --------------------------------------------------------------------------
    lex_lev = lex_data['chapter']
    txt_lev = corpus['text_lvl']['chapter']

    # --------------------------------------------------------------------------
    # Make the level label out of the simple name
    # --------------------------------------------------------------------------
    lvl_lbl = "chapters"
    lvl_lbl = lvl_lbl[:1].upper() + lvl_lbl[1:]

    # --------------------------------------------------------------------------
    # Header
    # --------------------------------------------------------------------------
    header = []
    header.append([lvl_lbl,"","","","","","","User"])
    header.append([])
    header.append(['Index','Chapter Idx','Page','Title','Num Paragraphs','Num Sentences',"x"])

    # --------------------------------------------------------------------------
    # Body
    # --------------------------------------------------------------------------
    body = []
    for i, idx in enumerate(txt_lev):
        # ----------------------------------------------------------------------
        # Get the lex data
        # ----------------------------------------------------------------------
        lexicon = lex_lev[idx]
        txt     = txt_lev[idx]

        # ----------------------------------------------------------------------
        # Tear apart the index
        # ----------------------------------------------------------------------
        cidx  = idx[0]
        page  = idx[1]
        title = idx[2]

        # ----------------------------------------------------------------------
        # Get the number of paragraphs and sentences
        # ----------------------------------------------------------------------
        pnum  = len(corpus['text'][idx]['paragraphs'])
        snum  = len(corpus['text'][idx]['sentences'])

        # ----------------------------------------------------------------------
        # Make the record and add it
        # ----------------------------------------------------------------------
        rec = []
        rec.append(i+1)
        rec.append(cidx)
        rec.append(page)
        rec.append(title)
        rec.append(pnum)
        rec.append(snum)

        body.append(rec)

    # --------------------------------------------------------------------------
    # Consolidate header and body
    # --------------------------------------------------------------------------
    data = matrixify([header,body])

    # --------------------------------------------------------------------------
    # Formatting
    # --------------------------------------------------------------------------
    fmt = {}
    fmt['column_widths']  = [12, 12, 12, 80, 12, 12, 5]
    fmt['reverse_rows']   = [1,2,3]
    fmt['short_rows']     = [2]
    fmt['wrap_rows']      = [3]
    fmt['fix_rows_after'] = 3
    fmt['freeze_panes']   = "A4"
    fmt['gutter_cols']    = ["G"]
    fmt['zoom']           = 80

    # --------------------------------------------------------------------------
    # Write the sheet in the file
    # --------------------------------------------------------------------------
    wb = xl.load_workbook(ifile)
    o = write_excel_sheet(wb, lvl_lbl, data, fmt)
    wb.save(ifile)
    wb.close()

    # --------------------------------------------------------------------------
    # Finish
    # --------------------------------------------------------------------------
    return True

# ------------------------------------------------------------------------------
# Write out one page to the spreadsheet for a particular 'level' of text
# ------------------------------------------------------------------------------
def write_output_page_paragraph(lex_data, ifile, corpus):
    # --------------------------------------------------------------------------
    # Start
    # --------------------------------------------------------------------------
    print("...writing paragraph data to spreadsheet...")

    # --------------------------------------------------------------------------
    # Get the lexicon data for this particular level
    # --------------------------------------------------------------------------
    lex_lev = lex_data['paragraph']
    txt_lev = corpus['text_lvl']['paragraph']
    dat_lev = corpus['data_lvl']['paragraph']

    # --------------------------------------------------------------------------
    # Make the level label out of the simple name
    # --------------------------------------------------------------------------
    lvl_lbl = "paragraphs"
    lvl_lbl = lvl_lbl[:1].upper() + lvl_lbl[1:]

    # --------------------------------------------------------------------------
    # Header
    # --------------------------------------------------------------------------
    header = []
    header.append([lvl_lbl,"","","","","","","User"])
    header.append([])
    header.append(['Index','Chapter Idx','Page','Title','Paragraph Idx' ,'Num Sentences',"x"])

    # --------------------------------------------------------------------------
    # Body
    # --------------------------------------------------------------------------
    body = []
    for i, idx in enumerate(txt_lev):
        # ----------------------------------------------------------------------
        # Get the lex data
        # ----------------------------------------------------------------------
        lexicon = lex_lev[idx]
        txt     = txt_lev[idx]
        data    = dat_lev[idx]

        # ----------------------------------------------------------------------
        # Tear apart the index
        # ----------------------------------------------------------------------
        cidx  = idx[0][0]
        page  = idx[0][1]
        title = idx[0][2]
        pidx  = idx[1]

        # ----------------------------------------------------------------------
        # Number of sentences for this paragraph
        # ----------------------------------------------------------------------
        snum = data['num_sentences_clean'] 

        # ----------------------------------------------------------------------
        # Make the record and add it
        # ----------------------------------------------------------------------
        rec = []
        rec.append(i+1)
        rec.append(cidx)
        rec.append(page)
        rec.append(title)
        rec.append(pidx)
        rec.append(snum)

        body.append(rec)

    # --------------------------------------------------------------------------
    # Consolidate header and body
    # --------------------------------------------------------------------------
    data = matrixify([header,body])

    # --------------------------------------------------------------------------
    # Formatting
    # --------------------------------------------------------------------------
    fmt = {}
    fmt['column_widths']  = [12, 12, 12, 80, 12, 12, 5]
    fmt['reverse_rows']   = [1,2,3]
    fmt['short_rows']     = [2]
    fmt['wrap_rows']      = [3]
    fmt['fix_rows_after'] = 3
    fmt['freeze_panes']   = "A4"
    fmt['gutter_cols']    = ["G"]
    fmt['zoom']           = 80

    # --------------------------------------------------------------------------
    # Write the sheet in the file
    # --------------------------------------------------------------------------
    wb = xl.load_workbook(ifile)
    o = write_excel_sheet(wb, lvl_lbl, data, fmt)
    wb.save(ifile)
    wb.close()

    # --------------------------------------------------------------------------
    # Finish
    # --------------------------------------------------------------------------
    return True

# ------------------------------------------------------------------------------
# Write out one page to the spreadsheet for a particular 'level' of text
# ------------------------------------------------------------------------------
def write_output_page_sentence(lex_data, ifile, corpus):
    # --------------------------------------------------------------------------
    # Start
    # --------------------------------------------------------------------------
    print("...writing sentence data to spreadsheet...")

    # --------------------------------------------------------------------------
    # Get the lexicon data for this particular level
    # --------------------------------------------------------------------------
    lex_lev = lex_data['sentence']
    txt_lev = corpus['text_lvl']['sentence']

    # --------------------------------------------------------------------------
    # Make the level label out of the simple name
    # --------------------------------------------------------------------------
    lvl_lbl = "sentences"
    lvl_lbl = lvl_lbl[:1].upper() + lvl_lbl[1:]


    # --------------------------------------------------------------------------
    # Header
    # --------------------------------------------------------------------------
    header = []
    header.append([lvl_lbl,"","","","","","","","","","User"])
    header.append([])
    header.append(['Index','Chapter Idx','Page','Chapter Title','Paragraph Idx' ,'Sentence Idx','Clean Sent Idx','Text','Cleaning Instruction','x'])
    
    # --------------------------------------------------------------------------
    # Body
    # --------------------------------------------------------------------------
    body = []
    for i, idx in enumerate(txt_lev):
        # ----------------------------------------------------------------------
        # Get the lex and txt data
        # ----------------------------------------------------------------------
        lexicon = lex_lev[idx]
        txt     = txt_lev[idx]

        # ----------------------------------------------------------------------
        # Tear apart the index
        # ----------------------------------------------------------------------
        cidx  = idx[0][0][0]
        page  = idx[0][0][1]
        title = idx[0][0][2]
        pidx  = idx[0][1]
        sidx  = idx[1]
        sidx1 = idx[2]

        # ----------------------------------------------------------------------
        # Get the cleaning
        # ----------------------------------------------------------------------
        cln = ""
        if cidx in corpus['cleaning']['text']:
            if pidx in corpus['cleaning']['text'][cidx]:
                if sidx in corpus['cleaning']['text'][cidx][pidx]:
                    cln = ";".join(corpus['cleaning']['text'][cidx][pidx][sidx])

        # ----------------------------------------------------------------------
        # Make the record and add it
        # ----------------------------------------------------------------------
        rec = []
        rec.append(i+1)
        rec.append(cidx)
        rec.append(page)
        rec.append(title)
        rec.append(pidx)
        rec.append(sidx)
        rec.append(sidx1)
        rec.append(txt)
        rec.append(cln)

        body.append(rec)

    # --------------------------------------------------------------------------
    # Consolidate header and body
    # --------------------------------------------------------------------------
    data = matrixify([header,body])

    # --------------------------------------------------------------------------
    # Formatting
    # --------------------------------------------------------------------------
    fmt = {}
    fmt['column_widths']        = [12, 12, 12, 40, 12, 12, 12, 130, 30, 5]
    fmt['reverse_rows']         = [1,2,3]
    fmt['short_rows']           = [2]
    fmt['wrap_rows']            = [3]
    fmt['wrap_cols']            = ["D","H"]
    fmt['fix_row_height_after'] = 3
    fmt['freeze_panes']         = "A4"
    fmt['gutter_cols']          = ["J"]
    fmt['zoom']                 = 80

    # --------------------------------------------------------------------------
    # Write the sheet in the file
    # --------------------------------------------------------------------------
    wb = xl.load_workbook(ifile)
    o = write_excel_sheet(wb, lvl_lbl, data, fmt)
    wb.save(ifile)
    wb.close()

    # --------------------------------------------------------------------------
    # Finish
    # --------------------------------------------------------------------------
    return True

# ------------------------------------------------------------------------------
# Put together the text files for the Obsidian vault
# ------------------------------------------------------------------------------
def compile_vault_output(corpus):
    # --------------------------------------------------------------------------
    # Start
    # --------------------------------------------------------------------------
    print("...compiling markup text...")

    # --------------------------------------------------------------------------
    # Compile the chapters as raw text
    # --------------------------------------------------------------------------
    chapters = {}
    for chapter in corpus['text']:
        # ----------------------------------------------------------------------
        # Make the chapter name
        # ----------------------------------------------------------------------
        cname = "CHAPTER " + str(chapter[0]) + ": " + chapter[2]

        # ----------------------------------------------------------------------
        # Start the output block of text 
        # ----------------------------------------------------------------------
        block = []

        # ----------------------------------------------------------------------
        # Chapter title
        # ----------------------------------------------------------------------
        block.append("# " + cname)
        block.append("")

        block.append("| | |")
        block.append("|------------|-----|")
        block.append("| Paragraphs |" + str(len(corpus['text'][chapter]['paragraphs'])) + "|")

        # ----------------------------------------------------------------------
        # Each paragraph
        # ----------------------------------------------------------------------
        for i, paragraph in enumerate(corpus['text'][chapter]['paragraphs']):
            # ------------------------------------------------------------------
            # Header
            # ------------------------------------------------------------------
            block.append("")
            block.append("---")
            block.append("## Paragraph " + str(i+1))
            block.append("")

            # ------------------------------------------------------------------
            # Text
            # ------------------------------------------------------------------
            block.append(corpus['text'][chapter]['paragraphs'][paragraph])

        # ----------------------------------------------------------------------
        # add the chapter
        # ----------------------------------------------------------------------
        chapters[cname] = block

    # --------------------------------------------------------------------------
    # Consolidate
    # --------------------------------------------------------------------------
    vault_data = {}
    vault_data['chapters'] = chapters

    # --------------------------------------------------------------------------
    # Finish
    # --------------------------------------------------------------------------
    return vault_data

# ------------------------------------------------------------------------------
# Write out the data to the text files
# ------------------------------------------------------------------------------
def write_vault_files(vault_data):
    # --------------------------------------------------------------------------
    # Start
    # --------------------------------------------------------------------------
    print("...writing text output...")
        
    # --------------------------------------------------------------------------
    # One file per chapter
    # --------------------------------------------------------------------------
    for cname in vault_data['chapters']:
        fpath = "vault\\chapters\\" + cname.split(":")[0] + ".md"
        with open(fpath, "w", encoding="utf-8") as f:
            for line in vault_data['chapters'][cname]:
                print(line, file=f)

    # --------------------------------------------------------------------------
    # Finish
    # --------------------------------------------------------------------------
    return True

# ------------------------------------------------------------------------------
# Run
# ------------------------------------------------------------------------------
if __name__ == "__main__":
    args = cli()
    start_time = datetime.utcnow()
    init = initialize(args)
    main(args, init)
    end_time     = datetime.utcnow()
    elapsed_time = end_time - start_time
    print("Elapsed time: " + str(elapsed_time))

